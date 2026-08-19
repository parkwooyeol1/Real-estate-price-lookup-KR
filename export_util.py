# -*- coding: utf-8 -*-
"""
조회 결과 내보내기 — 엑셀(.xlsx) / PDF(.pdf)

여러 주소의 공시가격·환산액(×1.3)·소유지분 반영액을 표로 저장한다.
- 엑셀: openpyxl
- PDF : reportlab (한글 폰트 '맑은 고딕' 등록; 없으면 다른 한글 폰트 자동 탐색)

의존성:  pip install openpyxl reportlab
"""

import os

MULT = 1.3  # 공시가격 환산 배율(앱과 동일)

# 표 헤더
COLUMNS = ["주소", "유형", "면적(㎡)", "기준연도",
           "공시가격(원)", "×1.3 환산(원)", "소유지분", "지분반영 최종(원)"]


def _bunui(share) -> str:
    if not share:
        return ""

    def t(x):
        return str(int(x)) if float(x).is_integer() else f"{x:g}"
    return f"{t(share['denominator'])}분의 {t(share['numerator'])}"


def iter_rows(results, shares) -> list[list]:
    """
    [(row_index, address, result_dict), ...] 와 {row_index: share} 로부터
    표의 각 행(list)을 만든다. 값이 없는(실패/미조회) 주소도 한 줄로 남긴다.
    """
    rows = []
    for idx, addr, r in (results or []):
        share = (shares or {}).get(idx)
        ratio = share["ratio"] if share else 1
        bunui = _bunui(share)
        disp_addr = addr or r.get("base_address") or ""

        if not r.get("ok"):
            rows.append([disp_addr, "조회 실패", "", "",
                         "", "", "", (r.get("message") or "").split("\n")[0]])
            continue
        if r.get("no_data"):
            rows.append([disp_addr, "미조회(신설·개편 행정구역 등)",
                         "", "", "", "", "", ""])
            continue

        sec = r.get("sections", {})
        items = []  # (유형, 면적, 연도, 공시가격 base)
        land = sec.get("land") or {}
        if land.get("jiga") and land.get("area"):
            items.append(("토지 공시가격", land.get("area"),
                          land.get("jiga_yeondo"), land["jiga"] * land["area"]))
        house = sec.get("house") or {}
        if house.get("price") is not None:
            items.append(("주택가격", house.get("lot_area"),
                          house.get("price_yeondo"), house["price"]))
        apt = sec.get("apt") or {}
        for row in (apt.get("rows") or []):
            if row.get("price") is None:
                continue
            dong = row.get("dong")
            dd = dong if (dong and str(dong).endswith("동")) else (f"{dong}동" if dong else "")
            unit = (f"{dd} " if dong else "") + f"{row.get('ho')}호"
            items.append((f"공동주택 {unit}".strip(), row.get("prvuse_area"),
                          row.get("price_yeondo"), row["price"]))

        if not items:
            rows.append([disp_addr, "환산 가격 없음", "", "", "", "", "", ""])
            continue

        for kind, ar, year, base in items:
            conv = base * MULT
            final = base * MULT * ratio
            rows.append([
                disp_addr, kind,
                (f"{ar:g}" if isinstance(ar, (int, float)) else ""),
                (str(year) if year else ""),
                int(round(base)), int(round(conv)),
                bunui, int(round(final)),
            ])
    return rows


# ── 엑셀 ─────────────────────────────
def export_excel(results, shares, path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "공시가격 조회결과"

    head_fill = PatternFill("solid", fgColor="1D1D1F")
    head_font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D2D2D7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    money_cols = {5, 6, 8}  # 공시가격/환산/최종
    for ri, row in enumerate(iter_rows(results, shares), start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="맑은 고딕")
            cell.border = border
            if ci in money_cols and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    widths = [34, 18, 10, 9, 16, 16, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    return path


# ── PDF ─────────────────────────────
_FONT_CANDIDATES = [
    r"C:\Windows\Fonts\malgun.ttf",              # 맑은 고딕 (Windows)
    r"C:\Windows\Fonts\malgunbd.ttf",
    "/System/Library/Fonts/AppleSDGothicNeo.ttc",  # macOS
    "/Library/Fonts/AppleGothic.ttf",
    "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",  # Linux
]


def _register_korean_font() -> str:
    """사용 가능한 한글 폰트를 reportlab 에 등록하고 폰트명을 반환."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    for p in _FONT_CANDIDATES:
        if os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("Korean", p))
                return "Korean"
            except Exception:
                continue
    return "Helvetica"  # 최후 폴백(한글은 깨질 수 있음)


def export_pdf(results, shares, path: str) -> str:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    font = _register_korean_font()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontName=font, fontSize=16)
    sub_style = ParagraphStyle("s", parent=styles["Normal"], fontName=font,
                               fontSize=9, textColor=colors.grey)

    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    story = [Paragraph("부동산 공시가격 조회 결과", title_style),
             Paragraph("공시가격 × 1.3 환산 (소유지분 입력 시 반영)", sub_style),
             Spacer(1, 6 * mm)]

    data = [COLUMNS]
    for row in iter_rows(results, shares):
        data.append([f"{v:,}" if isinstance(v, (int, float)) else v for v in row])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D1D1F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("ALIGN", (7, 1), (7, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D2D2D7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F7")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    doc.build(story)
    return path
